import os
import sys
import logging
import traceback
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
from werkzeug.middleware.proxy_fix import ProxyFix
import requests
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from concurrent.futures import ThreadPoolExecutor
import threading

from flask import session, redirect
from functools import wraps
import secrets as py_secrets

# === Paths absolutos a partir de __file__ (robusto a qualquer CWD) ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))                    # /app/backend
FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', 'frontend')) # /app/frontend

# === Logging que SEMPRE aparece no Railway ===
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s %(name)s: %(message)s',
    stream=sys.stdout,
    force=True,
)

# === Flask app: static_folder aponta para frontend INTEIRO ===
app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path='')

# Railway faz TLS termination; confia nos headers X-Forwarded-*
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

CORS(app)

# ── Configuração de sessão segura ────────────────────────────────────────────
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', py_secrets.token_hex(32))
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 horas

DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD', '')

def login_required(f):
    """Decorator que protege rotas - só acessa quem tem sessão válida."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            return jsonify({"error": "Não autorizado", "redirect": "/login"}), 401
        return f(*args, **kwargs)
    return decorated


# Liga logger do Flask ao do gunicorn (sem isso, tracebacks somem)
if __name__ != '__main__':
    gunicorn_logger = logging.getLogger('gunicorn.error')
    if gunicorn_logger.handlers:
        app.logger.handlers = gunicorn_logger.handlers
        app.logger.setLevel(gunicorn_logger.level)

# Log de startup pra confirmar que paths existem (aparece no Deploy Log)
app.logger.warning(f"BASE_DIR={BASE_DIR}")
app.logger.warning(f"FRONTEND_DIR={FRONTEND_DIR} exists={os.path.isdir(FRONTEND_DIR)}")
app.logger.warning(f"index.html exists={os.path.isfile(os.path.join(FRONTEND_DIR, 'index.html'))}")

# === Handler global: imprime traceback completo no stdout/stderr ===
@app.errorhandler(Exception)
def handle_any_exception(e):
    if isinstance(e, HTTPException):
        return e  # preserva 404, 405, etc
    tb = traceback.format_exc()
    print(f"\n=== UNHANDLED EXCEPTION on {request.method} {request.path} ===\n{tb}",
          file=sys.stderr, flush=True)
    app.logger.error("Unhandled exception on %s %s\n%s", request.method, request.path, tb)
    return jsonify(error=type(e).__name__, message=str(e)), 500

# === Rota de health pra validar deploy ===
@app.route('/api/health')
def health():
    return {
        "ok": True,
        "frontend_dir": FRONTEND_DIR,
        "index_exists": os.path.isfile(os.path.join(FRONTEND_DIR, 'index.html')),
    }

# ── Configurações ─────────────────────────────────────────────────────────────
CLIENT_ID     = os.getenv("ML_CLIENT_ID", "SEU_CLIENT_ID")
CLIENT_SECRET = os.getenv("ML_CLIENT_SECRET", "SEU_CLIENT_SECRET")
REDIRECT_URI  = os.getenv("ML_REDIRECT_URI", "http://localhost:5000/auth/callback")
TOKEN_FILE    = "tokens.json"
BASE_URL      = "https://api.mercadolibre.com"

# ── Cache de fretes (acelera consultas repetidas) ─────────────────────────────
SHIPMENT_CACHE_FILE = os.path.join(BASE_DIR, "shipments_cache.json")
_cache_lock = threading.Lock()

def load_shipment_cache():
    if os.path.exists(SHIPMENT_CACHE_FILE):
        try:
            with open(SHIPMENT_CACHE_FILE) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_shipment_cache(cache: dict):
    with _cache_lock:
        with open(SHIPMENT_CACHE_FILE, "w") as f:
            json.dump(cache, f)

def get_shipping_cost(ship_id, cache: dict):
    """Retorna custo de frete, usando cache quando possível."""
    key = str(ship_id)
    if key in cache:
        return cache[key]
    ship_data = ml_get(f"/shipments/{ship_id}")
    if ship_data:
        cost = float(ship_data.get("base_cost") or 0)
        cache[key] = cost
        return cost
    return 0

# ── Helpers de token ──────────────────────────────────────────────────────────
def save_tokens(data: dict):
    with open(TOKEN_FILE, "w") as f:
        json.dump(data, f)

def load_tokens():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE) as f:
            return json.load(f)
    return None

def refresh_access_token(refresh_token: str):
    r = requests.post(f"{BASE_URL}/oauth/token", data={
        "grant_type":    "refresh_token",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "refresh_token": refresh_token,
    })
    if r.ok:
        tokens = r.json()
        save_tokens(tokens)
        return tokens
    return None

def get_valid_token():
    tokens = load_tokens()
    if not tokens:
        return None
    return tokens.get("access_token")

def ml_get(path: str, params: dict = None):
    """GET autenticado na API do ML com auto-refresh."""
    tokens = load_tokens()
    if not tokens:
        return None

    headers = {"Authorization": f"Bearer {tokens['access_token']}"}
    r = requests.get(f"{BASE_URL}{path}", headers=headers, params=params or {})

    if r.status_code == 401:
        new_tokens = refresh_access_token(tokens.get("refresh_token", ""))
        if not new_tokens:
            return None
        headers["Authorization"] = f"Bearer {new_tokens['access_token']}"
        r = requests.get(f"{BASE_URL}{path}", headers=headers, params=params or {})

    return r.json() if r.ok else None

# ── Login do Dashboard ────────────────────────────────────────────────────────
@app.route("/api/dashboard-login", methods=["POST"])
def dashboard_login():
    """Valida senha de acesso ao dashboard."""
    data = request.get_json() or {}
    password = data.get("password", "")
    
    if not DASHBOARD_PASSWORD:
        return jsonify({"error": "Senha não configurada no servidor"}), 500
    
    if password == DASHBOARD_PASSWORD:
        session.permanent = True
        session['authenticated'] = True
        return jsonify({"success": True})
    
    return jsonify({"error": "Senha incorreta"}), 401

@app.route("/api/dashboard-logout", methods=["POST"])
def dashboard_logout():
    """Limpa sessão e desloga."""
    session.clear()
    return jsonify({"success": True})

@app.route("/api/dashboard-check")
def dashboard_check():
    """Verifica se a sessão tá válida."""
    return jsonify({"authenticated": bool(session.get('authenticated'))})

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/auth/login")
@login_required
def auth_login():
    url = (
        f"https://auth.mercadolivre.com.br/authorization"
        f"?response_type=code"
        f"&client_id={CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
    )
    return jsonify({"auth_url": url})

@app.route("/auth/callback")
def auth_callback():
    code = request.args.get("code")
    if not code:
        return "Erro: código não recebido.", 400

    r = requests.post(f"{BASE_URL}/oauth/token", data={
        "grant_type":    "authorization_code",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code":          code,
        "redirect_uri":  REDIRECT_URI,
    })

    if not r.ok:
        return f"Erro ao obter token: {r.text}", 400

    save_tokens(r.json())
    return """
    <html><body style="font-family:sans-serif;text-align:center;padding:60px">
      <h2>✅ Conta autorizada com sucesso!</h2>
      <p>Pode fechar esta aba e voltar ao dashboard.</p>
      <script>setTimeout(()=>window.close(),2000)</script>
    </body></html>
    """

@app.route("/auth/status")
@login_required
def auth_status():
    tokens = load_tokens()
    if not tokens:
        return jsonify({"authenticated": False})
    user = ml_get(f"/users/me")
    if not user:
        return jsonify({"authenticated": False})
    return jsonify({
        "authenticated": True,
        "user": {
            "id":       user.get("id"),
            "nickname": user.get("nickname"),
            "email":    user.get("email"),
        }
    })

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/api/dashboard")
@login_required
def dashboard():
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")

    # Padrão: hoje
    if not date_from:
        date_from = datetime.now().strftime("%Y-%m-%dT00:00:00.000-03:00")
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%dT23:59:59.000-03:00")

    user_data = ml_get("/users/me")
    if not user_data:
        return jsonify({"error": "Não autenticado"}), 401

    seller_id = user_data["id"]

    # ── PAGINAÇÃO: busca TODOS os pedidos do período ──
    all_orders = []
    offset = 0
    page_size = 50
    max_pages = 100  # até 5000 pedidos por consulta
    total = 0

    for _ in range(max_pages):
        orders_data = ml_get("/orders/search", {
            "seller": seller_id,
            "order.date_created.from": date_from,
            "order.date_created.to":   date_to,
            "sort":   "date_desc",
            "limit":  page_size,
            "offset": offset,
        })

        if not orders_data:
            return jsonify({"error": "Erro ao buscar pedidos"}), 500

        results = orders_data.get("results", [])
        total = orders_data.get("paging", {}).get("total", 0)
        all_orders.extend(results)

        if len(results) < page_size or len(all_orders) >= total:
            break

        offset += page_size

    orders = all_orders

    # ── BUSCA FRETES EM PARALELO (com cache) ──
    cache = load_shipment_cache()
    shipping_costs = {}  # order_id -> custo

    def fetch_one(order):
        oid = order.get("id")
        shipment = order.get("shipping") or {}
        ship_id = shipment.get("id")
        if not ship_id:
            return oid, 0
        cost = get_shipping_cost(ship_id, cache)
        return oid, cost

    # Filtra só pedidos não-cancelados pra poupar chamadas
    fetch_list = [o for o in orders if o.get("status") != "cancelled"]

    with ThreadPoolExecutor(max_workers=15) as executor:
        for oid, cost in executor.map(fetch_one, fetch_list):
            shipping_costs[oid] = cost

    # Salva cache atualizado de uma vez
    save_shipment_cache(cache)

    # ── Agrega métricas ──
    total_revenue   = 0.0
    total_shipping  = 0.0
    paid_orders     = 0
    pending_orders  = 0
    cancelled_orders= 0
    items_sold      = 0
    order_list      = []

    for o in orders:
        status = o.get("status", "")
        if status == "cancelled":
            cancelled_orders += 1
            continue

        total_amount  = o.get("total_amount", 0) or 0
        shipping_cost = shipping_costs.get(o.get("id"), 0)

        if status == "paid":
            paid_orders    += 1
            total_revenue  += total_amount
            total_shipping += shipping_cost
        elif status in ("confirmed", "payment_required", "payment_in_process"):
            pending_orders += 1

        for item in o.get("order_items", []):
            items_sold += item.get("quantity", 0)

        order_list.append({
            "id":            o.get("id"),
            "date":          o.get("date_created", "")[:10],
            "status":        status,
            "buyer":         (o.get("buyer") or {}).get("nickname", "—"),
            "item":          (o.get("order_items") or [{}])[0].get("item", {}).get("title", "—"),
            "quantity":      sum(i.get("quantity", 0) for i in o.get("order_items", [])),
            "total":         total_amount,
            "shipping_cost": shipping_cost,
        })

    return jsonify({
        "summary": {
            "total_orders":    total,
            "paid_orders":     paid_orders,
            "pending_orders":  pending_orders,
            "cancelled_orders":cancelled_orders,
            "items_sold":      items_sold,
            "total_revenue":   round(total_revenue, 2),
            "total_shipping":  round(total_shipping, 2),
            "net_revenue":     round(total_revenue - total_shipping, 2),
        },
        "orders": order_list,
        "period": {"from": date_from[:10], "to": date_to[:10]},
    })

# ── Exportar Excel ────────────────────────────────────────────────────────────
@app.route("/api/export/excel")
@login_required
def export_excel():
    date_from = request.args.get("date_from", datetime.now().strftime("%Y-%m-%dT00:00:00.000-03:00"))
    date_to   = request.args.get("date_to",   datetime.now().strftime("%Y-%m-%dT23:59:59.000-03:00"))

    # Reutiliza a lógica do dashboard
    with app.test_request_context(f"/api/dashboard?date_from={date_from}&date_to={date_to}"):
        resp = dashboard()
        data = resp.get_json() if hasattr(resp, "get_json") else json.loads(resp[0].data)

    if "error" in data:
        return jsonify(data), 401

    wb = Workbook()

    # ── Aba Resumo ────────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resumo"

    yellow  = PatternFill("solid", fgColor="FFD700")
    green   = PatternFill("solid", fgColor="1E8449")
    dark    = PatternFill("solid", fgColor="1A1A2E")
    white_f = Font(color="FFFFFF", bold=True, size=12)
    black_f = Font(bold=True, size=11)

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = f"📊 Relatório Mercado Livre — {data['period']['from']} a {data['period']['to']}"
    ws_sum["A1"].font    = Font(bold=True, size=14, color="FFFFFF")
    ws_sum["A1"].fill    = dark
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    headers = ["Métrica", "Valor"]
    ws_sum.append([])
    ws_sum.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=3, column=col, value=h)
        cell.font = white_f
        cell.fill = green
        cell.alignment = Alignment(horizontal="center")

    s = data["summary"]
    rows = [
        ("Total de Pedidos",        s["total_orders"]),
        ("Pedidos Pagos",           s["paid_orders"]),
        ("Pedidos Pendentes",       s["pending_orders"]),
        ("Pedidos Cancelados",      s["cancelled_orders"]),
        ("Itens Vendidos",          s["items_sold"]),
        ("Faturamento Bruto",       f"R$ {s['total_revenue']:,.2f}"),
        ("Total Frete",             f"R$ {s['total_shipping']:,.2f}"),
        ("Faturamento Líquido",     f"R$ {s['net_revenue']:,.2f}"),
    ]
    for r in rows:
        ws_sum.append(list(r))

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 22

    # ── Aba Pedidos ───────────────────────────────────────────────────────────
    ws_ord = wb.create_sheet("Pedidos")
    ord_headers = ["ID Pedido", "Data", "Status", "Comprador", "Produto", "Qtd", "Total (R$)", "Frete (R$)"]
    ws_ord.append(ord_headers)
    for col, h in enumerate(ord_headers, 1):
        cell = ws_ord.cell(row=1, column=col, value=h)
        cell.font = white_f
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center")

    STATUS_PT = {
        "paid": "Pago", "confirmed": "Confirmado",
        "payment_required": "Aguard. Pagamento",
        "payment_in_process": "Pagamento em processamento",
        "cancelled": "Cancelado",
    }
    light = PatternFill("solid", fgColor="F2F2F2")
    for i, o in enumerate(data["orders"]):
        row = [
            o["id"], o["date"],
            STATUS_PT.get(o["status"], o["status"]),
            o["buyer"], o["item"],
            o["quantity"], o["total"], o["shipping_cost"],
        ]
        ws_ord.append(row)
        if i % 2 == 1:
            for col in range(1, 9):
                ws_ord.cell(row=i+2, column=col).fill = light

    for col in range(1, 9):
        ws_ord.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_ml_{data['period']['from']}_{data['period']['to']}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Catch-all: serve frontend (DEVE VIR POR ÚLTIMO!) ─────────────────────────
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    full = os.path.join(FRONTEND_DIR, path)
    if path and os.path.isfile(full):
        return send_from_directory(FRONTEND_DIR, path)
    return send_from_directory(FRONTEND_DIR, 'index.html')


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
